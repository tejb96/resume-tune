"""Job-application tracker spreadsheet: append, load, and update."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HEADERS: tuple[str, ...] = (
    "Date Applied",
    "Company",
    "Role / Job Title",
    "Location",
    "Status",
    "Job URL",
    "ATS Score at Apply",
    "Follow-up Date",
    "ATS Keywords",
    "Resume File",
    "JD Snapshot",
    "Notes",
)

_COLUMN_WIDTHS: dict[str, float] = {
    "Date Applied": 14,
    "Company": 22,
    "Role / Job Title": 28,
    "Location": 18,
    "Status": 14,
    "Job URL": 30,
    "ATS Score at Apply": 18,
    "Follow-up Date": 14,
    "ATS Keywords": 25,
    "Resume File": 40,
    "JD Snapshot": 40,
    "Notes": 35,
}


def _cell_to_value(val: Any) -> Any:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.date().isoformat() if val.time() == datetime.min.time() else val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    return val


def _read_header_row(ws) -> list[str]:
    headers: list[str] = []
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is None or str(val).strip() == "":
            break
        headers.append(str(val).strip())
    return headers


def _apply_header_formatting(ws) -> None:
    bold = Font(bold=True)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _COLUMN_WIDTHS[header]
    ws.freeze_panes = "A2"


def _load_rows_from_worksheet(ws) -> list[dict[str, Any]]:
    file_headers = _read_header_row(ws)
    if not file_headers:
        return []

    rows: list[dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        if all(ws.cell(row=row_idx, column=col).value in (None, "") for col in range(1, len(file_headers) + 1)):
            continue
        row_dict = {header: "" for header in HEADERS}
        for col_idx, header in enumerate(file_headers, start=1):
            if header not in HEADERS:
                continue
            row_dict[header] = _cell_to_value(ws.cell(row=row_idx, column=col_idx).value)
        rows.append(row_dict)
    return rows


def ensure_tracker(tracker_path: Path) -> None:
    """Create the tracker spreadsheet with headers and formatting if it does not exist."""
    if tracker_path.is_file():
        return

    tracker_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    _apply_header_formatting(ws)
    wb.save(tracker_path)


def migrate_tracker(tracker_path: Path) -> None:
    """Rewrite the tracker to the canonical HEADERS layout, preserving row data by header name."""
    ensure_tracker(tracker_path)
    wb = load_workbook(tracker_path)
    ws = wb.active
    current_headers = _read_header_row(ws)
    if current_headers == list(HEADERS):
        wb.close()
        return
    rows = _load_rows_from_worksheet(ws)
    wb.close()
    save_applications(tracker_path, rows)


def load_applications(tracker_path: Path) -> list[dict[str, Any]]:
    """Load all application rows keyed by canonical HEADERS."""
    migrate_tracker(tracker_path)
    wb = load_workbook(tracker_path)
    ws = wb.active
    rows = _load_rows_from_worksheet(ws)
    wb.close()
    return rows


def save_applications(tracker_path: Path, rows: list[dict[str, Any]]) -> None:
    """Replace all data rows in the tracker, preserving header formatting."""
    tracker_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    _apply_header_formatting(ws)

    for row in rows:
        ws.append([row.get(header, "") for header in HEADERS])

    wb.save(tracker_path)


def update_application_row(tracker_path: Path, row_index: int, row: dict[str, Any]) -> None:
    """Update one application row by zero-based index in the loaded row list."""
    rows = load_applications(tracker_path)
    if row_index < 0 or row_index >= len(rows):
        raise IndexError(f"Application row index out of range: {row_index}")
    rows[row_index] = {header: row.get(header, "") for header in HEADERS}
    save_applications(tracker_path, rows)


def log_application(tracker_path: Path, row: dict) -> None:
    """Append one application row to the tracker spreadsheet."""
    migrate_tracker(tracker_path)
    wb = load_workbook(tracker_path)
    ws = wb.active
    ws.append([row.get(header, "") for header in HEADERS])
    wb.save(tracker_path)
