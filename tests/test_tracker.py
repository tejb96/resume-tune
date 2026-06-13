"""Tests for application tracker spreadsheet helpers."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from resume_tune.tracker.tracker import (
    HEADERS,
    ensure_tracker,
    load_applications,
    log_application,
    migrate_tracker,
    save_applications,
)

LEGACY_HEADERS: tuple[str, ...] = (
    "Date Applied",
    "Company",
    "Role / Job Title",
    "Location",
    "Status",
    "ATS Keywords",
    "Resume File",
    "Notes",
)


def _legacy_row() -> dict:
    return {
        "Date Applied": "2026-06-11",
        "Company": "Acme Corp",
        "Role / Job Title": "Software Engineer",
        "Location": "Remote",
        "Status": "Applied",
        "ATS Keywords": "Python, AWS",
        "Resume File": "/tmp/resume.docx",
        "Notes": "Referred by friend",
    }


def _full_row(**overrides) -> dict:
    row = {
        "Date Applied": "2026-06-11",
        "Company": "Acme Corp",
        "Role / Job Title": "Software Engineer",
        "Location": "Remote",
        "Status": "Applied",
        "Job URL": "https://jobs.example.com/123",
        "ATS Score at Apply": 82.5,
        "Follow-up Date": "2026-06-18",
        "ATS Keywords": "Python, AWS",
        "Resume File": "/tmp/resume.docx",
        "JD Snapshot": "/tmp/jd_snapshots/2026-06-11_Acme.txt",
        "Notes": "Referred by friend",
    }
    row.update(overrides)
    return row


def _write_legacy_workbook(tracker_path: Path, row: dict) -> None:
    tracker_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(list(LEGACY_HEADERS))
    ws.append([row.get(header, "") for header in LEGACY_HEADERS])
    wb.save(tracker_path)


def test_ensure_tracker_creates_formatted_workbook(tmp_path: Path) -> None:
    tracker_path = tmp_path / "applications.xlsx"
    ensure_tracker(tracker_path)

    assert tracker_path.is_file()

    wb = load_workbook(tracker_path)
    ws = wb.active
    assert [ws.cell(row=1, column=i).value for i in range(1, len(HEADERS) + 1)] == list(
        HEADERS
    )
    assert all(ws.cell(row=1, column=i).font.bold for i in range(1, len(HEADERS) + 1))
    assert ws.freeze_panes == "A2"
    assert ws.column_dimensions["A"].width == 14
    assert ws.column_dimensions["B"].width == 22


def test_ensure_tracker_is_idempotent(tmp_path: Path) -> None:
    tracker_path = tmp_path / "applications.xlsx"
    ensure_tracker(tracker_path)
    ensure_tracker(tracker_path)
    assert tracker_path.is_file()
    wb = load_workbook(tracker_path)
    assert wb.active.max_row == 1


def test_log_application_appends_row(tmp_path: Path) -> None:
    tracker_path = tmp_path / "applications.xlsx"
    row = _full_row()
    log_application(tracker_path, row)

    wb = load_workbook(tracker_path)
    ws = wb.active
    assert ws.max_row == 2
    assert [ws.cell(row=2, column=i).value for i in range(1, len(HEADERS) + 1)] == [
        row[header] for header in HEADERS
    ]
    assert all(ws.cell(row=1, column=i).font.bold for i in range(1, len(HEADERS) + 1))


def test_log_application_appends_second_row(tmp_path: Path) -> None:
    tracker_path = tmp_path / "applications.xlsx"
    base_row = _full_row()
    log_application(tracker_path, base_row)
    second = {**base_row, "Company": "Beta Inc", "Resume File": "/tmp/resume2.docx"}
    log_application(tracker_path, second)

    wb = load_workbook(tracker_path)
    ws = wb.active
    assert ws.max_row == 3
    assert ws.cell(row=3, column=2).value == "Beta Inc"
    assert ws.cell(row=3, column=10).value == "/tmp/resume2.docx"


def test_migrate_tracker_adds_new_columns_for_legacy_workbook(tmp_path: Path) -> None:
    tracker_path = tmp_path / "applications.xlsx"
    legacy_row = _legacy_row()
    _write_legacy_workbook(tracker_path, legacy_row)

    migrate_tracker(tracker_path)

    wb = load_workbook(tracker_path)
    ws = wb.active
    assert [ws.cell(row=1, column=i).value for i in range(1, len(HEADERS) + 1)] == list(
        HEADERS
    )
    loaded = load_applications(tracker_path)
    assert len(loaded) == 1
    assert loaded[0]["Company"] == legacy_row["Company"]
    assert loaded[0]["Resume File"] == legacy_row["Resume File"]
    assert loaded[0]["Job URL"] == ""
    assert loaded[0]["JD Snapshot"] == ""


def test_load_and_save_applications_round_trip(tmp_path: Path) -> None:
    tracker_path = tmp_path / "applications.xlsx"
    rows = [
        _full_row(),
        _full_row(Company="Beta Inc", Status="Interview"),
    ]
    save_applications(tracker_path, rows)

    loaded = load_applications(tracker_path)
    assert loaded == rows

    loaded[1]["Notes"] = "Updated note"
    save_applications(tracker_path, loaded)
    reloaded = load_applications(tracker_path)
    assert reloaded[1]["Notes"] == "Updated note"
