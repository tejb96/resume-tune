"""Append job-application rows to a local Excel tracker spreadsheet."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HEADERS: tuple[str, ...] = (
    "Date Applied",
    "Company",
    "Role / Job Title",
    "Location",
    "Status",
    "ATS Keywords",
    "Resume File",
    "Notes",
)

_COLUMN_WIDTHS: dict[str, float] = {
    "Date Applied": 14,
    "Company": 22,
    "Role / Job Title": 28,
    "Location": 18,
    "Status": 14,
    "ATS Keywords": 25,
    "Resume File": 40,
    "Notes": 35,
}


def ensure_tracker(tracker_path: Path) -> None:
    """Create the tracker spreadsheet with headers and formatting if it does not exist."""
    if tracker_path.is_file():
        return

    tracker_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    bold = Font(bold=True)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _COLUMN_WIDTHS[header]

    ws.freeze_panes = "A2"
    wb.save(tracker_path)


def log_application(tracker_path: Path, row: dict) -> None:
    """Append one application row to the tracker spreadsheet."""
    ensure_tracker(tracker_path)
    wb = load_workbook(tracker_path)
    ws = wb.active
    ws.append([row.get(header, "") for header in HEADERS])
    wb.save(tracker_path)
